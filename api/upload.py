"""
文件上传 API 路由
"""

from fastapi import APIRouter, UploadFile, File, HTTPException
from pathlib import Path
import uuid
import shutil
import os

from services.db_service import DBService

router = APIRouter()
db = DBService()

# 上传目录
BASE_DIR = Path(__file__).parent.parent
UPLOAD_DIR = BASE_DIR / "uploads"
PDF_DIR = UPLOAD_DIR / "pdf"
EXCEL_DIR = UPLOAD_DIR / "excel"

# 确保目录存在
PDF_DIR.mkdir(parents=True, exist_ok=True)
EXCEL_DIR.mkdir(parents=True, exist_ok=True)


@router.post("/pdf")
async def upload_pdf(file: UploadFile = File(...)):
    """
    上传 PDF 报告表

    Returns:
        file_id: 文件唯一标识
        filename: 文件名
        page_count: 页数
    """
    # 检查文件类型
    if not (file.filename and file.filename.lower().endswith(".pdf")):
        raise HTTPException(status_code=400, detail="只支持 PDF 文件")

    # 生成唯一文件 ID
    file_id = str(uuid.uuid4())
    file_path = PDF_DIR / f"{file_id}.pdf"

    # 保存文件
    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"文件保存失败: {str(e)}")

    # 获取页数
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(file_path)
        page_count = len(doc)
        doc.close()
    except Exception:
        page_count = 0  # 无法获取页数

    return {
        "success": True,
        "file_id": file_id,
        "filename": file.filename,
        "page_count": page_count,
        "message": f"上传成功，共 {page_count} 页"
    }


@router.post("/excel")
async def upload_excel(file: UploadFile = File(...)):
    """
    上传 Excel 反馈文件

    Returns:
        file_id: 文件唯一标识
        filename: 文件名
        sheets: 工作表列表
    """
    # 检查文件类型
    if not (file.filename and file.filename.lower().endswith((".xlsx", ".xls"))):
        raise HTTPException(status_code=400, detail="只支持 Excel 文件 (.xlsx, .xls)")

    # 生成唯一文件 ID
    file_id = str(uuid.uuid4())
    safe_name = file.filename if file.filename else "unnamed"
    file_path = EXCEL_DIR / f"{file_id}_{safe_name}"

    # 保存文件
    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"文件保存失败: {str(e)}")

    # 保存到数据库
    sheets = []
    row_counts = {}
    try:
        import openpyxl
        wb_temp = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        sheets = list(wb_temp.sheetnames)
        for sn in sheets:
            ws_temp = wb_temp[sn]
            row_counts[sn] = ws_temp.max_row or 0
        wb_temp.close()
        db.save_feedback_file(file_id, safe_name, sheets, row_counts)
        # 按章节保存
        wb_data = openpyxl.load_workbook(str(file_path), data_only=True)
        for sn in wb_data.sheetnames:
            ws_data = wb_data[sn]
            rows = list(ws_data.iter_rows(values_only=True))
            if rows:
                headers = [str(c) if c else "" for c in rows[0]]
                data_rows = [dict(zip(headers, [str(c) if c else "" for c in row])) for row in rows[1:] if any(row)]
                if data_rows:
                    db.save_feedback_chapter(file_id, sn, data_rows)
        wb_data.close()
    except Exception as e:
        sheets = []  # 出错时返回空列表，不影响文件上传

    return {
        "success": True,
        "file_id": file_id,
        "filename": file.filename,
        "sheets": sheets,
        "sheet_count": len(sheets),
        "message": f"上传成功，共 {len(sheets)} 个工作表"
    }


@router.get("/files")
async def list_files():
    """
    列出已上传的文件
    """
    pdf_files = []
    excel_files = []

    # 列出 PDF 文件
    for f in PDF_DIR.glob("*.pdf"):
        pdf_files.append({
            "id": f.stem,
            "filename": f.name,
            "size": f.stat().st_size,
        })

    # 列出 Excel 文件（同时从 DB 获取元数据）
    db_files = {f["file_id"]: f for f in db.list_feedback_files(limit=100)}
    for f in EXCEL_DIR.glob("*.xlsx"):
        file_id = f.stem.split("_")[0] if "_" in f.stem else f.stem
        excel_files.append({
            "id": file_id,
            "filename": f.name,
            "size": f.stat().st_size,
            "db": db_files.get(file_id, {}),
        })

    return {
        "pdf_files": pdf_files,
        "excel_files": excel_files,
    }


@router.get("/history")
async def get_history():
    """
    获取历史记录摘要（来自数据库）
    """
    summary = db.get_summary()
    # 获取最近 20 条任务
    recent_tasks = db.list_scan_tasks(limit=20)
    recent_compares = db.list_compare_results(limit=20)
    recent_feedback = db.list_feedback_files(limit=20)
    return {
        "summary": summary,
        "recent_tasks": recent_tasks,
        "recent_compares": recent_compares,
        "recent_feedback": recent_feedback,
    }


@router.get("/pdf/{file_id}")
async def get_pdf(file_id: str):
        """获取已上传的 PDF 文件（供前端预览）"""
        for f in PDF_DIR.glob(f"{file_id}.pdf"):
                from fastapi.responses import FileResponse
                return FileResponse(f, media_type="application/pdf", filename=f.name)
        raise HTTPException(status_code=404, detail="文件不存在")


@router.delete("/{file_type}/{file_id}")
async def delete_file(file_type: str, file_id: str):
        """
        删除已上传的文件
        """
        if file_type == "pdf":
                target_dir = PDF_DIR
                pattern = f"{file_id}.pdf"
        elif file_type == "excel":
                target_dir = EXCEL_DIR
                pattern = f"{file_id}_*"
        else:
                raise HTTPException(status_code=400, detail="无效的文件类型")

        # 查找并删除文件
        deleted = False
        for f in target_dir.glob(pattern):
                f.unlink()
                deleted = True

        if deleted:
                return {"success": True, "message": "文件已删除"}
        else:
                raise HTTPException(status_code=404, detail="文件不存在")
