"""
Excel 处理服务
读取查核反馈 Excel 文件
"""

import openpyxl
from pathlib import Path
from typing import Dict, List, Optional, Any
from datetime import datetime


class ExcelService:
    """Excel 处理服务"""

    # Sheet 名称与章节的映射关系
    SHEET_MAPPING = {
        "房产": "境内房产",
        "房产明细": "境内房产",
        "境外房产": "境外房产",
        "保险": "境内投资型保险",
        "境外保险": "境外投资型保险",
        "基金": "境内基金",
        "境外基金": "境外基金",
        "护照1": "本人护照",
        "护照2": "本人护照",
        "通行证": "港澳台通行证",
        "出入境": "因私出国",
        "法人": "投资企业",
        "企业主要管理人员": "投资企业",
        "股东": "投资企业",
    }

    def __init__(self):
        pass

    def read_excel(self, excel_path: str | Path, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        读取 Excel 文件

        Args:
            excel_path: Excel 文件路径
            sheet_name: 工作表名称，None 表示全部

        Returns:
            包含各工作表数据的字典
        """
        excel_path = Path(excel_path)

        if not excel_path.exists():
            raise FileNotFoundError(f"Excel 文件不存在: {excel_path}")

        wb = openpyxl.load_workbook(excel_path, data_only=True)

        if sheet_name:
            # 读取指定工作表
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            return {
                sheet_name: self._read_sheet(wb[sheet_name])
            }
        else:
            # 读取全部工作表
            result = {}
            for name in wb.sheetnames:
                result[name] = self._read_sheet(wb[name])
            return result

    def _read_sheet(self, sheet) -> Dict[str, Any]:
        """
        读取单个工作表

        Args:
            sheet: openpyxl 工作表对象

        Returns:
            工作表数据
        """
        data = {
            "headers": [],
            "rows": [],
            "row_count": 0,
            "col_count": 0
        }

        if sheet.max_row < 1:
            return data

        # 读取表头（第一行）
        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value) if cell.value is not None else "")
        data["headers"] = headers
        data["col_count"] = len(headers)

        # 读取数据行
        for row_idx in range(2, sheet.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                row_data[header] = self._format_cell_value(cell.value)
            data["rows"].append(row_data)

        data["row_count"] = len(data["rows"])
        return data

    def _format_cell_value(self, value: Any) -> Any:
        """格式化单元格值"""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, float):
            # 判断是否为整数
            if value == int(value):
                return int(value)
            return round(value, 2)
        return value

    def get_sheet_data(self, excel_path: str | Path, sheet_name: str) -> List[Dict]:
        """
        获取指定工作表的数据行

        Args:
            excel_path: Excel 文件路径
            sheet_name: 工作表名称

        Returns:
            数据行列表，每行为字典
        """
        result = self.read_excel(excel_path, sheet_name)
        if sheet_name in result:
            return result[sheet_name].get("rows", [])
        return []

    def get_all_feedback_data(self, excel_path: str | Path) -> Dict[str, List[Dict]]:
        """
        获取所有反馈数据，按映射章节分组

        Args:
            excel_path: Excel 文件路径

        Returns:
            按章节分组的数据
        """
        all_data = self.read_excel(excel_path)

        result = {}
        for sheet_name, sheet_data in all_data.items():
            # 映射到对应章节
            chapter = self.SHEET_MAPPING.get(sheet_name, sheet_name)
            if chapter not in result:
                result[chapter] = []
            result[chapter].extend(sheet_data.get("rows", []))

        return result

    def get_summary(self, excel_path: str | Path) -> Dict[str, Any]:
        """
        获取 Excel 文件摘要

        Args:
            excel_path: Excel 文件路径

        Returns:
            摘要信息
        """
        all_data = self.read_excel(excel_path)

        summary = {
            "total_sheets": len(all_data),
            "sheets": [],
            "total_rows": 0
        }

        for sheet_name, sheet_data in all_data.items():
            sheet_info = {
                "name": sheet_name,
                "mapped_to": self.SHEET_MAPPING.get(sheet_name, sheet_name),
                "row_count": sheet_data.get("row_count", 0),
                "col_count": sheet_data.get("col_count", 0),
                "headers": sheet_data.get("headers", [])
            }
            summary["sheets"].append(sheet_info)
            summary["total_rows"] += sheet_info["row_count"]

        return summary
