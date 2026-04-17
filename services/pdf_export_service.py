"""
PDF 导出服务
将 OCR 识别结果、反馈数据、比對结果导出为表单式 PDF
"""

import json
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime
from io import BytesIO

from fpdf import FPDF
from openpyxl import load_workbook


class PDFExportService:
    """PDF 导出服务"""

    # 报告表章节顺序和标题
    CHAPTER_ORDER = [
        ("封面", ["报告人签名", "单位", "报告日期"]),
        ("本人基本情况", ["姓名", "性别", "民族", "政治面貌", "身份证号码", "JG证号码", "现住址"]),
        ("本人婚姻情况", ["婚姻现状", "婚姻变化", "以往变化情形及时间"]),
        ("本人健康状况", ["是否身患重大疾病", "疾病名称", "确诊时间", "诊断医疗机构", "治疗康复情况"]),
        ("本人护照情况", ["护照号码", "签发日期", "有效期至", "保管机构"]),
        ("本人因私出国", ["起止日期_起", "起止日期_止", "所到国家", "事由", "审批机构"]),
        ("港澳台通行证", ["证件名称", "证件号码", "有效期限_起", "有效期限_止", "保管机构"]),
        ("因私往来港澳台", ["起止日期_起", "起止日期_止", "所到地区", "事由", "审批机构"]),
        ("配偶基本情况", ["配偶姓名", "配偶工作单位", "配偶现任职务", "配偶单位性质", "配偶政治面貌", "配偶身份证号码"]),
        ("子女从业情况", ["子女姓名", "子女关系", "是否共同生活", "子女工作单位", "子女现任职务", "子女单位性质", "子女身份证号码"]),
        ("境内房产", ["产权人", "房产来源", "具体地址", "建筑面积", "房产性质", "交易时间_年", "交易价格"]),
        ("境外房产", ["产权人", "来源", "国家地区", "具体地址", "面积", "购买时间", "交易价格", "币种"]),
        ("境内股票", ["持有人", "股票名称", "股票代码", "持股数量", "市值(万元)"]),
        ("境外股票", ["持有人", "股票名称", "股票代码", "国家地区", "持股数量", "市值(万元)", "币种"]),
        ("境内基金", ["持有人", "基金名称", "基金代码", "基金份额", "净值(万元)"]),
        ("境外基金", ["持有人", "基金名称", "基金代码", "国家地区", "基金份额", "净值(万元)", "币种"]),
        ("境内投资型保险", ["投保人", "保险产品", "保单号", "保险公司", "累计保费(万元)"]),
        ("境外投资型保险", ["投保人", "保险产品", "保单号", "国家地区", "保险公司", "币种", "累计保费(万元)"]),
        ("投资企业", ["姓名", "企业名称", "信用代码", "企业类型", "注册资本(万元)", "是否股东", "认缴出资额(万元)", "认缴比例"]),
        ("配偶子女开办中介机构", ["姓名", "资格名称", "执业证号", "机构名称", "机构类型", "注册资本", "是否从业"]),
        ("私募股权基金", ["姓名", "基金产品名称", "基金编码", "实缴金额(万元)", "基金管理人"]),
        ("境外存款", ["存款人", "开户银行", "账号", "币种", "金额(万元)"]),
        ("其他事项", ["其他事项内容", "本人承诺签名"]),
    ]

    # 字段名中英文映射
    FIELD_ALIASES = {
        "姓名": "name",
        "地址": "address",
        "建筑面积": "area",
        "面积": "area",
        "房产性质": "property_type",
        "交易时间_年": "year",
        "交易时间_月": "month",
        "交易价格": "price",
        "交易价格(万元)": "price",
        "备注": "remark",
        "保单号": "policy_no",
        "护照号码": "passport_no",
        "基金名称": "fund_name",
        "基金代码": "fund_code",
        "企业名称": "company_name",
        "信用代码": "credit_code",
        "注册资本": "registered_capital",
        "注册资本(万元)": "registered_capital",
        "是否股东": "is_shareholder",
        "认缴出资额": "subscribed_amount",
        "认缴出资额(万元)": "subscribed_amount",
        "认缴比例": "ratio",
        "认缴比例(%)": "ratio",
    }

    def __init__(self):
        self.pdf = None

    def _make_pdf(self) -> FPDF:
        """创建 PDF 实例"""
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        font_dir = str(Path(__file__).parent.parent / "fonts")
        pdf.add_font("Chinese", "", f"{font_dir}/msyh.ttc", uni=True)
        pdf.add_font("Chinese", "B", f"{font_dir}/msyhbd.ttc", uni=True)
        return pdf

    def export_scan_result(self, ocr_result_file: str | Path, output_file: str | Path) -> Path:
        """
        导出 OCR 识别结果为表单式 PDF

        Args:
            ocr_result_file: OCR 结果 JSON 文件路径
            output_file: 输出 PDF 文件路径

        Returns:
            输出文件路径
        """
        with open(ocr_result_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        pdf = self._make_pdf()
        pdf.add_page()

        # 标题
        pdf.set_font("Chinese", "B", 16)
        pdf.cell(0, 12, "报告表 OCR 识别结果", ln=True, align="C")
        pdf.ln(5)

        # 元信息
        pdf.set_font("Chinese", "", 10)
        pdf.cell(0, 7, f"识别时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
        pdf.cell(0, 7, f"源文件: {Path(ocr_result_file).name}", ln=True)
        pdf.ln(5)

        # 统计
        total_pages = len(data)
        total_regions = sum(len(p.get("regions", [])) for p in data)
        pdf.set_font("Chinese", "B", 11)
        pdf.set_fill_color(240, 240, 240)
        pdf.cell(60, 8, f"总页数: {total_pages}", fill=True)
        pdf.cell(60, 8, f"差异区域数: {total_regions}", fill=True)
        pdf.ln(12)

        # 按页展示
        for page_data in data:
            page_num = page_data.get("page", "?")
            regions = page_data.get("regions", [])
            diff_ratio = page_data.get("diff_ratio", 0)

            pdf.set_font("Chinese", "B", 12)
            pdf.set_fill_color(220, 230, 240)
            pdf.cell(0, 9, f"第 {page_num} 页  (差异率: {diff_ratio:.2f}%)", fill=True, ln=True)

            if not regions:
                pdf.set_font("Chinese", "", 10)
                pdf.cell(0, 7, "  无差异区域", ln=True)
            else:
                for region in regions:
                    bbox = region.get("bbox", [])
                    ocr_result = region.get("ocr_result", {})
                    texts = ocr_result.get("texts", [])
                    tables = ocr_result.get("tables", [])

                    # 区域信息
                    pdf.set_font("Chinese", "", 9)
                    pdf.set_fill_color(248, 248, 248)
                    pdf.cell(0, 6, f"  区域 {region.get('region_index', '?')}: bbox={bbox}", fill=True, ln=True)

                    # 文本内容
                    if texts:
                        for text_item in texts:
                            content = text_item.get("content", "")
                            label = text_item.get("label", "text")
                            if content and content.strip():
                                pdf.set_font("Chinese", "", 9)
                                # 截断过长的内容
                                if len(content) > 60:
                                    content = content[:60] + "..."
                                pdf.cell(0, 5, f"    [{label}] {content}", ln=True)

                    # 表格内容
                    if tables:
                        for table_item in tables:
                            content = table_item.get("content", "")
                            if content:
                                pdf.set_font("Chinese", "", 9)
                                # 简化 HTML 表格
                                import re
                                content_clean = re.sub(r"<[^>]+>", " ", content)
                                content_clean = re.sub(r"\s+", " ", content_clean).strip()
                                if len(content_clean) > 80:
                                    content_clean = content_clean[:80] + "..."
                                pdf.cell(0, 5, f"    [表格] {content_clean}", ln=True)

            pdf.ln(3)

        pdf.output(str(output_file))
        return Path(output_file)

    def export_feedback(self, excel_file: str | Path, output_file: str | Path) -> Path:
        """
        导出反馈 Excel 数据为表单式 PDF

        Args:
            excel_file: Excel 文件路径
            output_file: 输出 PDF 文件路径

        Returns:
            输出文件路径
        """
        wb = load_workbook(str(excel_file), data_only=True)

        pdf = self._make_pdf()
        pdf.add_page()

        # 标题
        pdf.set_font("Chinese", "B", 16)
        pdf.cell(0, 12, "查核反馈信息", ln=True, align="C")
        pdf.ln(3)

        # 元信息
        pdf.set_font("Chinese", "", 10)
        pdf.cell(0, 7, f"文件: {Path(excel_file).name}", ln=True)
        pdf.cell(0, 7, f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
        pdf.ln(5)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 1:
                continue

            # 分页
            pdf.add_page()
            pdf.set_font("Chinese", "B", 14)
            pdf.set_fill_color(60, 90, 150)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(0, 10, f"  {sheet_name}", fill=True, ln=True)
            pdf.set_text_color(0, 0, 0)
            pdf.ln(3)

            # 表头
            headers = [str(cell.value) if cell.value else "" for cell in ws[1]]
            col_widths = self._compute_col_widths(ws, headers)
            pdf.set_font("Chinese", "B", 9)
            pdf.set_fill_color(200, 220, 240)
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 7, header[:20], border=1, fill=True)
            pdf.ln()

            # 数据行
            pdf.set_font("Chinese", "", 9)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                # 交替行颜色
                if row_idx % 2 == 0:
                    pdf.set_fill_color(248, 248, 248)
                else:
                    pdf.set_fill_color(255, 255, 255)

                for col_idx, value in enumerate(row):
                    if col_idx >= len(col_widths):
                        break
                    cell_text = str(value) if value is not None else ""
                    if len(cell_text) > 25:
                        cell_text = cell_text[:25] + ".."
                    pdf.cell(col_widths[col_idx], 6, cell_text, border=1, fill=True)
                pdf.ln()

        pdf.output(str(output_file))
        return Path(output_file)

    def export_compare_result(self, compare_result_file: str | Path, output_file: str | Path) -> Path:
        """
        导出比对结果为表单式 PDF

        Args:
            compare_result_file: 比对结果 JSON 文件路径
            output_file: 输出 PDF 文件路径

        Returns:
            输出文件路径
        """
        with open(compare_result_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        pdf = self._make_pdf()
        pdf.add_page()

        # 标题
        pdf.set_font("Chinese", "B", 16)
        pdf.set_text_color(180, 0, 0)
        pdf.cell(0, 12, "领导干部个人有关事项报告表 核查比对结果", ln=True, align="C")
        pdf.set_text_color(0, 0, 0)
        pdf.ln(3)

        # 摘要
        summary = data.get("summary", {})
        stats = summary.get("statistics", {})
        pdf.set_font("Chinese", "B", 11)
        pdf.cell(0, 7, f"比对时间: {summary.get('compare_time', '')}", ln=True)
        pdf.cell(0, 7, f"比对ID: {summary.get('compare_id', '')}", ln=True)
        pdf.ln(3)

        # 统计卡片
        pdf.set_font("Chinese", "B", 10)
        cards = [
            ("比对总数", stats.get("total", 0), (100, 100, 100)),
            ("一致", stats.get("consistent", 0), (50, 150, 50)),
            ("不一致", stats.get("inconsistent", 0), (200, 50, 50)),
            ("缺失", stats.get("missing", 0), (200, 150, 0)),
        ]
        for label, value, color in cards:
            pdf.set_fill_color(*color)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(45, 10, f"{label}: {value}", fill=True)
            pdf.set_text_color(0, 0, 0)
        pdf.ln(14)

        # 一致率
        pdf.set_font("Chinese", "B", 12)
        rate = summary.get("consistency_rate", "N/A")
        pdf.cell(0, 8, f"一致率: {rate}", ln=True)
        pdf.ln(5)

        # 差异项列表
        diff_items = data.get("diff_items", [])

        # 分组
        inconsistent = [d for d in diff_items if d.get("result") == "不一致"]
        missing = [d for d in diff_items if d.get("result") == "缺失"]

        if inconsistent:
            pdf.add_page()
            pdf.set_font("Chinese", "B", 13)
            pdf.set_fill_color(200, 50, 50)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(0, 9, f"  不一致项目 ({len(inconsistent)} 项)", fill=True, ln=True)
            pdf.set_text_color(0, 0, 0)
            pdf.ln(2)

            self._render_diff_table(pdf, inconsistent)

        if missing:
            pdf.add_page()
            pdf.set_font("Chinese", "B", 13)
            pdf.set_fill_color(200, 150, 0)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(0, 9, f"  缺失项目 ({len(missing)} 项)", fill=True, ln=True)
            pdf.set_text_color(0, 0, 0)
            pdf.ln(2)

            self._render_diff_table(pdf, missing)

        if not diff_items:
            pdf.set_font("Chinese", "", 11)
            pdf.cell(0, 10, "所有项目均一致，无差异项", ln=True, align="C")

        pdf.output(str(output_file))
        return Path(output_file)

    def _render_diff_table(self, pdf: FPDF, items: List[Dict]):
        """渲染差异项表格"""
        pdf.set_font("Chinese", "B", 9)
        pdf.set_fill_color(230, 230, 230)
        headers = ["章节", "序号", "字段", "报告表填写", "反馈信息", "状态"]
        widths = [25, 10, 25, 50, 50, 20]
        for i, h in enumerate(headers):
            pdf.cell(widths[i], 7, h, border=1, fill=True)
        pdf.ln()

        pdf.set_font("Chinese", "", 8)
        for idx, item in enumerate(items):
            fill = (idx % 2 == 0)
            pdf.set_fill_color(255, 245, 245) if fill else pdf.set_fill_color(255, 255, 255)

            def cell(text, w, max_chars=20):
                t = str(text) if text else ""
                if len(t) > max_chars:
                    t = t[:max_chars] + ".."
                pdf.cell(w, 6, t, border=1, fill=fill)

            cell(item.get("chapter", ""), 25)
            cell(item.get("index", ""), 10)
            cell(item.get("field", ""), 25)
            cell(item.get("report_value", ""), 50)
            cell(item.get("feedback_value", ""), 50)
            cell(item.get("result", ""), 20)
            pdf.ln()

        pdf.ln(3)

    def _compute_col_widths(self, ws, headers: List[str]) -> List[float]:
        """计算列宽"""
        available = 190  # A4 宽度减去边距
        n = len(headers)
        base = available / n
        return [base] * n
