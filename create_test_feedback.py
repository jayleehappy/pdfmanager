"""
创建测试反馈数据 Excel 文件
用于测试比对功能
"""

import openpyxl
from pathlib import Path
import uuid

BASE_DIR = Path(__file__).parent
FEEDBACK_DIR = BASE_DIR / "uploads/excel"
FEEDBACK_DIR.mkdir(parents=True, exist_ok=True)


def create_feedback_excel():
    """创建测试反馈数据 Excel"""
    wb = openpyxl.Workbook()

    # 清理默认 sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # === Sheet 1: 房产 ===
    ws1 = wb.create_sheet("房产")
    headers1 = ["姓名", "地址", "建筑面积(平方米)", "房产性质", "交易时间_年", "交易时间_月", "交易价格(万元)", "备注"]
    ws1.append(headers1)
    # 测试场景1: 完全一致
    ws1.append(["张三", "北京市朝阳区建国路88号1号楼101室", 120.5, "商品房", 2020, 6, 850.0, ""])
    # 测试场景2: 面积不一致
    ws1.append(["李四", "上海市浦东新区陆家嘴环路1000号", 250.0, "商品房", 2019, 3, 1200.0, ""])
    # 测试场景3: 报告表缺失此房产
    ws1.append(["王五", "广州市天河区珠江新城花城大道98号", 180.0, "商品房", 2021, 9, 600.0, "报告表无此房产"])

    # === Sheet 2: 保险 ===
    ws2 = wb.create_sheet("保险")
    headers2 = ["姓名", "保险公司名称", "保险产品全称", "保单号", "累计缴纳保费(万元)", "备注"]
    ws2.append(headers2)
    ws2.append(["张三", "中国人寿保险股份有限公司", "国寿福瑞安康保险", "GZ202000123456", 50.0, ""])
    ws2.append(["李四", "中国平安保险股份有限公司", "平安福满分保险", "PA202100987654", 30.0, "保单号不同"])

    # === Sheet 3: 基金 ===
    ws3 = wb.create_sheet("基金")
    headers3 = ["姓名", "基金名称", "基金代码", "基金份额", "填报前一交易日净值(万元)", "备注"]
    ws3.append(headers3)
    ws3.append(["张三", "易方达消费行业股票A", "110022", 50000, 1.85, ""])
    ws3.append(["李四", "华夏沪深300ETF联接A", "000051", 30000, 4.12, "份额不同"])

    # === Sheet 4: 护照1 ===
    ws4 = wb.create_sheet("护照1")
    headers4 = ["姓名", "护照号码", "签发日期", "有效期至", "保管机构", "备注"]
    ws4.append(headers4)
    ws4.append(["张三", "E12345678", "2020-01-15", "2030-01-15", "本人保管", ""])
    ws4.append(["李四", "E87654321", "2019-06-20", "2029-06-20", "单位保管", "报告表无此护照"])

    # === Sheet 5: 法人（投资企业） ===
    ws5 = wb.create_sheet("法人")
    headers5 = ["姓名", "统一社会信用代码", "企业名称", "企业类型", "注册资本(万元)", "是否股东", "个人认缴出资额(万元)", "认缴出资比例(%)"]
    ws5.append(headers5)
    ws5.append(["张三", "91110000MA00ABCD01", "北京创新科技有限公司", "有限责任公司", 1000.0, "是", 300.0, 30.0, ""])
    ws5.append(["李四", "91310000MB00EFGH02", "上海智慧科技有限公司", "股份有限公司", 5000.0, "是", 1000.0, 20.0, "企业名称相似但比例不同"])

    # === Sheet 6: 出入境 ===
    ws6 = wb.create_sheet("出入境")
    headers6 = ["姓名", "起止日期_起", "起止日期_止", "所到国家", "事由", "审批机构", "所用护照号"]
    ws6.append(headers6)
    ws6.append(["张三", "2023-07-01", "2023-07-15", "日本", "旅游", "北京市公安局", "E12345678"])

    # 设置列宽
    for ws in [ws1, ws2, ws3, ws4, ws5, ws6]:
        for col_idx, header in enumerate(ws[1], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    # 保存
    file_id = str(uuid.uuid4())[:8]
    filename = f"test_feedback_{file_id}.xlsx"
    filepath = FEEDBACK_DIR / filename
    wb.save(filepath)
    print(f"[OK] 反馈数据已保存: {filepath}")
    print(f"[OK] file_id: {file_id}")
    return file_id, filepath


if __name__ == "__main__":
    file_id, filepath = create_feedback_excel()
    print(f"\n上传此文件到系统后，可使用以下 file_id 进行比对测试:")
    print(f"  curl -X POST http://localhost:8000/api/upload/excel -F 'file=@{filepath}'")
